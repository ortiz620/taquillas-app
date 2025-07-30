# Requiere: Python + Flask + Google Sheets

from flask import Flask, render_template, request, redirect, url_for, send_file
from datetime import datetime
import os
import json
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from io import BytesIO

app = Flask(__name__)

taquillas = []
cortes = {}
gastos = {}

# Conexión a Google Sheets
def connect_to_sheets():
    credentials_json = os.environ.get("GOOGLE_SHEETS_CREDENTIALS")
    sheet_id = os.environ.get("GOOGLE_SPREADSHEET_ID")
    if not credentials_json or not sheet_id:
        return None

    credentials_dict = json.loads(credentials_json)
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    creds = ServiceAccountCredentials.from_json_keyfile_dict(credentials_dict, scope)
    client = gspread.authorize(creds)
    sheet = client.open_by_key(sheet_id)
    return sheet

@app.route('/')
def index():
    ingreso_total = sum(t['total'] for t in taquillas)
    gasto_total = sum(sum(g['monto'] for g in lista) for lista in gastos.values())
    neto_total = ingreso_total - gasto_total
    
    return render_template('index.html', 
                         taquillas=taquillas, 
                         cortes=cortes, 
                         gastos=gastos,
                         ingreso_total=ingreso_total,
                         gasto_total=gasto_total,
                         neto_total=neto_total)

@app.route('/agregar_taquilla', methods=['POST'])
def agregar_taquilla():
    nombre = request.form['nombre']
    precio = float(request.form['precio'])
    taquillas.append({'nombre': nombre, 'precio': precio, 'inicial': 0, 'final': 0, 'total': 0})
    return redirect(url_for('index'))

@app.route('/actualizar_boletos', methods=['POST'])
def actualizar_boletos():
    for i, t in enumerate(taquillas):
        inicial = int(request.form.get(f'inicial_{i}', 0))
        final = int(request.form.get(f'final_{i}', 0))
        vendidos = max(0, final - inicial - 2)
        total = vendidos * t['precio']
        taquillas[i]['inicial'] = inicial
        taquillas[i]['final'] = final
        taquillas[i]['total'] = total
    return redirect(url_for('index'))

@app.route('/agregar_gasto', methods=['POST'])
def agregar_gasto():
    descripcion = request.form['descripcion']
    monto = float(request.form['monto'])
    hoy = datetime.now().strftime('%Y-%m-%d')
    if hoy not in gastos:
        gastos[hoy] = []
    gastos[hoy].append({'descripcion': descripcion, 'monto': monto})
    return redirect(url_for('index'))

@app.route('/guardar_google_sheets')
def guardar_google_sheets():
    sheet = connect_to_sheets()
    if not sheet:
        return "Google Sheets no configurado"

    hoy = datetime.now().strftime('%Y-%m-%d')
    try:
        worksheet = sheet.worksheet(hoy)
    except:
        worksheet = sheet.add_worksheet(title=hoy, rows="100", cols="10")

    worksheet.clear()

    # Sección: Taquillas
    worksheet.append_row(["--------------------------------------"])
    worksheet.append_row(["TAQUILLAS DEL DÍA"])
    worksheet.append_row(["--------------------------------------"])
    worksheet.append_row(["Taquilla", "Inicial", "Final", "Precio", "Total"])
    for t in taquillas:
        worksheet.append_row([t['nombre'], t['inicial'], t['final'], t['precio'], t['total']])
    
    worksheet.append_row([""])  # Espacio vacío

    # Sección: Gastos
    worksheet.append_row(["--------------------------------------"])
    worksheet.append_row(["GASTOS DEL DÍA"])
    worksheet.append_row(["--------------------------------------"])
    worksheet.append_row(["Descripción", "Monto"])
    for g in gastos.get(hoy, []):
        worksheet.append_row([g['descripcion'], g['monto']])

    worksheet.append_row([""])  # Espacio vacío

    # Sección: Resumen
    ingreso_total = sum(t['total'] for t in taquillas)
    gasto_total = sum(g['monto'] for g in gastos.get(hoy, []))
    neto = ingreso_total - gasto_total

    worksheet.append_row(["--------------------------------------"])
    worksheet.append_row(["RESUMEN DEL DÍA"])
    worksheet.append_row(["--------------------------------------"])
    worksheet.append_row(["Ingreso Bruto", ingreso_total])
    worksheet.append_row(["Gasto Bruto", gasto_total])
    worksheet.append_row(["Ingreso Neto", neto])

    return redirect(url_for('index'))

@app.route('/descargar_excel')
def descargar_excel():
    # Crear un nuevo workbook de Excel
    wb = Workbook()
    ws = wb.active
    hoy = datetime.now().strftime('%Y-%m-%d')
    ws.title = f"Corte_{hoy}"
    
    # Configurar estilos
    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center')
    
    row = 1
    
    # Sección: Taquillas del Día
    ws[f'A{row}'] = "--------------------------------------"
    row += 1
    ws[f'A{row}'] = "TAQUILLAS DEL DÍA"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].alignment = center_alignment
    row += 1
    ws[f'A{row}'] = "--------------------------------------"
    row += 1
    
    # Headers para taquillas
    headers = ['Taquilla', 'Inicial', 'Final', 'Precio', 'Total']
    for col, header in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=header).font = header_font
    row += 1
    
    # Datos de taquillas
    for t in taquillas:
        ws.cell(row=row, column=1, value=t['nombre'])
        ws.cell(row=row, column=2, value=t['inicial'])
        ws.cell(row=row, column=3, value=t['final'])
        ws.cell(row=row, column=4, value=t['precio'])
        ws.cell(row=row, column=5, value=t['total'])
        row += 1
    
    # Espacio vacío
    row += 1
    
    # Sección: Gastos del Día
    ws[f'A{row}'] = "--------------------------------------"
    row += 1
    ws[f'A{row}'] = "GASTOS DEL DÍA"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].alignment = center_alignment
    row += 1
    ws[f'A{row}'] = "--------------------------------------"
    row += 1
    
    # Headers para gastos
    ws.cell(row=row, column=1, value='Descripción').font = header_font
    ws.cell(row=row, column=2, value='Monto').font = header_font
    row += 1
    
    # Datos de gastos del día actual
    for g in gastos.get(hoy, []):
        ws.cell(row=row, column=1, value=g['descripcion'])
        ws.cell(row=row, column=2, value=g['monto'])
        row += 1
    
    # Espacio vacío
    row += 1
    
    # Sección: Resumen del Día
    ws[f'A{row}'] = "--------------------------------------"
    row += 1
    ws[f'A{row}'] = "RESUMEN DEL DÍA"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].alignment = center_alignment
    row += 1
    ws[f'A{row}'] = "--------------------------------------"
    row += 1
    
    # Calcular totales
    ingreso_total = sum(t['total'] for t in taquillas)
    gasto_total = sum(g['monto'] for g in gastos.get(hoy, []))
    neto = ingreso_total - gasto_total
    
    # Datos del resumen
    ws.cell(row=row, column=1, value='Ingreso Bruto').font = header_font
    ws.cell(row=row, column=2, value=ingreso_total)
    row += 1
    ws.cell(row=row, column=1, value='Gasto Bruto').font = header_font
    ws.cell(row=row, column=2, value=gasto_total)
    row += 1
    ws.cell(row=row, column=1, value='Ingreso Neto').font = header_font
    ws.cell(row=row, column=2, value=neto)
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Corte_{hoy}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

if __name__ == '__main__':
    app.run(debug=True)
